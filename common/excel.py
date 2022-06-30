# _*_ coding:utf-8 _*-
# @author：尹乐
import openpyxl
class PlayExcel():
    def __init__(self,filname,sheet):
        # 选择表单
        self.w = openpyxl.load_workbook(filname) #读取文件
        self.sh= self.w[sheet]



    #读所有内容,用列表加元祖显示
    def read_all(self):

        # 选择单元格对象
        max_column = self.sh.max_column
        max_row = self.sh.max_row

        l1 = []
        for j in range(1, max_row + 1):
            l = []
            for i in range(1, max_column + 1):
                table = self.sh.cell(row=j, column=i)
                l.append(table.value)
                # print(table.value,end=" ")

            l1.append(tuple(l))
            # print("")

        return l1
    # 写入excel
    def writer_result(self):
        pass