import openpyxl
from common import excel_write

from common import excel_write
from common import get_local
import xlwt
# #打开一个表格
# work=openpyxl.Workbook()
# #创建一个表单
# work.create_sheet("test")
# #保存
# work.save("1.xlsx")


# a=excel.PlayExcel("1.xlsx","test").read_all()
# print(a)

if __name__ == '__main__':
    # wb = openpyxl.Workbook()
    filepath=get_local.get_cwd()
    #
    # ws=wb.active
    # ws.title="test"
    list=[1,2,4,5,6,7,34,54,234,45,97,56,2,35]
    # for i in list:
    #     ws.append([i])
    # wb.save(filepath + r"\3.xlsx")

    excel_write.Write_excel(filepath,r"\1234.xlsx","test",list).write_result()


